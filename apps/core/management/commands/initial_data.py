"""
Django Management Command to populate initial data for Everest Cakes

Usage:
    python manage.py initial_data

This command will create:
- Site Settings
- Product Categories
- Product Attributes & Options
- Product Addons
- Products with Variants
- Testimonials
- Hero Sections
- Featured Cards
- CMS Pages
"""

from django.core.management.base import BaseCommand
from django.core.files import File
from django.conf import settings
from django.db import transaction
from django.utils.text import slugify
from apps.core.models import (
    SiteSetting, Page, HeroSection, FeaturedCard, Testimonial
)
from apps.products.models import (
    Category, Product, ProductVariant, ProductAttribute, 
    ProductAttributeOption, ProductAttributeMapping, ProductAddon, ProductReview
)
import os
import requests
from io import BytesIO
from decimal import Decimal
from pathlib import Path
import re

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class Command(BaseCommand):
    help = 'Populate initial data for Everest Cakes e-commerce platform'

    def add_arguments(self, parser):
        parser.add_argument(
            '--reset',
            action='store_true',
            help='Delete existing data before creating new data',
        )
        parser.add_argument(
            '--catalog-file',
            type=str,
            help='Path to Excel (.xlsx) file containing catalog sheets for categories, addons, products, and variants',
        )

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('🚀 Starting initial data population...'))

        catalog_file = options.get('catalog_file')

        if catalog_file:
            if options['reset']:
                self.stdout.write('⚠️  Reset flag enabled - clearing existing catalog data...')
                self.clear_catalog_data()
            self.import_catalog_from_excel(catalog_file)
            self.stdout.write(self.style.SUCCESS('\n✅ Catalog import from Excel complete!'))
            self.stdout.write(self.style.SUCCESS('📊 Catalog Summary:'))
            self.stdout.write(f'   - Categories: {Category.objects.count()}')
            self.stdout.write(f'   - Products: {Product.objects.count()}')
            self.stdout.write(f'   - Variants: {ProductVariant.objects.count()}')
            self.stdout.write(f'   - Addons: {ProductAddon.objects.count()}')
            return
        
        if options['reset']:
            self.stdout.write('⚠️  Reset flag enabled - clearing existing data...')
            self.clear_data()
        
        # Create all data
        self.create_site_settings()
        self.create_cms_pages()
        self.create_categories()
        self.create_attributes()
        self.create_addons()
        self.create_products()
        self.create_testimonials()
        self.create_hero_sections()
        self.create_featured_cards()
        
        self.stdout.write(self.style.SUCCESS('\n✅ Initial data population complete!'))
        self.stdout.write(self.style.SUCCESS('📊 Summary:'))
        self.stdout.write(f'   - Categories: {Category.objects.count()}')
        self.stdout.write(f'   - Products: {Product.objects.count()}')
        self.stdout.write(f'   - Attributes: {ProductAttribute.objects.count()}')
        self.stdout.write(f'   - Addons: {ProductAddon.objects.count()}')
        self.stdout.write(f'   - Testimonials: {Testimonial.objects.count()}')

    def clear_catalog_data(self):
        """Clear existing catalog data only (products and related entities)."""
        debug_original = settings.DEBUG
        try:
            settings.DEBUG = False
            from django.db import connection

            delete_tables = [
                'cart_cartitem',
                'products_productreview',
                'products_productvariant',
                'products_productattributemapping_available_options',
                'products_productattributemapping',
                'products_product_addons',
                'products_product_tags',
                'products_product',
                'products_productattributeoption',
                'products_productattribute',
                'products_productaddon',
                'products_category',
            ]

            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                existing_tables = {row[0] for row in cursor.fetchall()}
                for table in delete_tables:
                    if table in existing_tables:
                        cursor.execute(f'DELETE FROM "{table}"')
        finally:
            settings.DEBUG = debug_original

        self.stdout.write('   ✓ Catalog data cleared')

    def import_catalog_from_excel(self, file_path):
        """Import categories, addons, products, variants and product-addon mappings from Excel."""
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for Excel import. Install dependencies from requirements.txt first.')

        excel_path = Path(file_path)
        if not excel_path.exists() or not excel_path.is_file():
            raise RuntimeError(f'Excel file not found: {excel_path}')

        if excel_path.suffix.lower() != '.xlsx':
            raise RuntimeError('Only .xlsx files are supported for --catalog-file')

        self.stdout.write(f'📥 Loading catalog workbook: {excel_path}')
        workbook = load_workbook(filename=excel_path, data_only=True)

        matrix_sheet = self._detect_price_matrix_sheet(workbook)
        if matrix_sheet is not None:
            self.stdout.write(f"📊 Detected price-matrix format in sheet '{matrix_sheet.title}'")
            self._import_price_matrix_catalog(matrix_sheet)
            return

        categories_sheet = self._get_sheet_by_aliases(workbook, ['categories', 'category'])
        addons_sheet = self._get_sheet_by_aliases(workbook, ['addons', 'addon', 'product_addons'])
        products_sheet = self._get_sheet_by_aliases(workbook, ['products', 'product'])
        variants_sheet = self._get_sheet_by_aliases(workbook, ['variants', 'variant', 'product_variants'])
        product_addons_sheet = self._get_sheet_by_aliases(workbook, ['product_addons', 'productaddons', 'product_addon_map'], required=False)

        self.stdout.write('📁 Importing categories...')
        category_count, categories_by_slug, categories_by_name = self._import_categories_sheet(categories_sheet)
        self.stdout.write(f'   ✓ Categories imported: {category_count}')

        self.stdout.write('➕ Importing addons...')
        addon_count, addons_by_slug, addons_by_name = self._import_addons_sheet(addons_sheet)
        self.stdout.write(f'   ✓ Addons imported: {addon_count}')

        self.stdout.write('🎂 Importing products...')
        products_by_slug, products_by_name = self._import_products_sheet(
            products_sheet,
            categories_by_slug,
            categories_by_name,
        )
        self.stdout.write(f'   ✓ Products imported: {len(products_by_slug)}')

        self.stdout.write('📦 Importing variants...')
        variant_count = self._import_variants_sheet(variants_sheet, products_by_slug, products_by_name)
        self.stdout.write(f'   ✓ Variants imported: {variant_count}')

        addon_link_count = self._map_product_addons_from_products_sheet(
            products_sheet,
            products_by_slug,
            products_by_name,
            addons_by_slug,
            addons_by_name,
        )

        if product_addons_sheet is not None:
            addon_link_count += self._import_product_addons_sheet(
                product_addons_sheet,
                products_by_slug,
                products_by_name,
                addons_by_slug,
                addons_by_name,
            )

        self.stdout.write(f'   ✓ Product-addon links imported: {addon_link_count}')

    def _normalize_header(self, value):
        if value is None:
            return ''
        text = str(value).strip().lower()
        normalized = ''.join(ch if ch.isalnum() else '_' for ch in text)
        while '__' in normalized:
            normalized = normalized.replace('__', '_')
        return normalized.strip('_')

    def _normalize_key(self, value):
        if value is None:
            return ''
        text = str(value).strip().lower()
        return ' '.join(text.split())

    def _slug(self, value):
        return slugify(str(value).strip()) if value is not None else ''

    def _parse_decimal(self, value, default=Decimal('0')):
        if value in (None, ''):
            return default
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace(',', '')
        if text == '':
            return default
        try:
            return Decimal(text)
        except Exception:
            return default

    def _parse_int(self, value, default=0):
        if value in (None, ''):
            return default
        try:
            return int(float(value))
        except Exception:
            return default

    def _parse_bool(self, value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().lower()
        if text in {'1', 'true', 'yes', 'y', 'on'}:
            return True
        if text in {'0', 'false', 'no', 'n', 'off'}:
            return False
        return default

    def _sheet_rows(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        raw_headers = rows[0]
        headers = [self._normalize_header(cell) for cell in raw_headers]
        data_rows = []
        for row in rows[1:]:
            if not row or all(cell in (None, '') for cell in row):
                continue
            row_dict = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row_dict[header] = row[index] if index < len(row) else None
            data_rows.append(row_dict)
        return headers, data_rows

    def _detect_price_matrix_sheet(self, workbook):
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers, _ = self._sheet_rows(sheet)
            if 'name' not in headers:
                continue
            if any(re.search(r'\d+(?:\.\d+)?\s*kg$', header or '', re.IGNORECASE) for header in headers):
                return sheet
        return None

    def _extract_note_addons(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        normalized_rows = [
            [str(cell).strip() if cell is not None else '' for cell in row]
            for row in rows
        ]

        addons = []

        # Hard icing / fondant note
        for row in normalized_rows:
            for cell in row:
                lower = cell.lower()
                if 'hard icing' in lower or 'fondant' in lower:
                    price_match = re.search(r'(\d{2,6})', lower)
                    price = Decimal(price_match.group(1)) if price_match else Decimal('0')
                    addons.append({
                        'name': 'Hard Icing / Fondant Upgrade',
                        'slug': 'hard-icing-fondant-upgrade',
                        'description': cell,
                        'price': price,
                        'is_free': price == 0,
                        'max_quantity': 1,
                        'order': 1,
                        'is_available': True,
                    })
                    break

        edible_col = None
        topper_col = None
        edible_row = None
        topper_row = None

        for row_index, row in enumerate(normalized_rows):
            for col_index, cell in enumerate(row):
                lower = cell.lower()
                if edible_col is None and 'edible' in lower and 'photo' in lower:
                    edible_col = col_index
                    edible_row = row_index
                if topper_col is None and 'paper topper' in lower:
                    topper_col = col_index
                    topper_row = row_index

        def add_size_addons(start_row, col_index, prefix, order_start):
            if col_index is None or start_row is None:
                return []
            parsed = []
            order = order_start
            for row_index in range(start_row + 1, min(start_row + 10, len(normalized_rows))):
                cell = normalized_rows[row_index][col_index] if col_index < len(normalized_rows[row_index]) else ''
                if not cell:
                    continue
                size_match = re.search(r'\b(A[0-9]+)\b', cell, re.IGNORECASE)
                price_match = re.search(r'(\d{2,6})', cell)
                if not size_match or not price_match:
                    continue
                size = size_match.group(1).upper()
                price = Decimal(price_match.group(1))
                name = f'{prefix} {size}'
                parsed.append({
                    'name': name,
                    'slug': self._slug(name),
                    'description': f'{prefix} size {size}',
                    'price': price,
                    'is_free': price == 0,
                    'max_quantity': 2,
                    'order': order,
                    'is_available': True,
                })
                order += 1
            return parsed

        addons.extend(add_size_addons(edible_row, edible_col, 'Edible Photo', 10))
        addons.extend(add_size_addons(topper_row, topper_col, 'Paper Topper', 20))

        deduped = {}
        for addon in addons:
            deduped[addon['slug']] = addon
        return list(deduped.values())

    def _normalized_matrix_title(self, raw_name):
        text = ' '.join(str(raw_name or '').strip().split())
        key = text.lower()

        aliases = {
            'vannila': 'Vanilla Cakes',
            'strawberry': 'Strawberry Cakes',
            'chocolate': 'Chocolate Cakes',
            'marble': 'Marble Cakes',
            'orange': 'Orange Cakes',
            'lemon': 'Lemon Cakes',
            'blackforest': 'Black Forest Cakes',
            'passion': 'Passion Cakes',
            'white forest': 'White Forest Cakes',
            'butter scotch': 'Butterscotch Cakes',
            'redvelvet cake': 'Red Velvet Cakes',
            'eggless cake': 'Eggless Cakes',
            'diabetic cake': 'Diabetic Cakes',
            'cinnamon cherry': 'Cinnamon Cherry Cakes',
            'carrot pineapple': 'Carrot Pineapple Cakes',
            'carrot nut cake': 'Carrot Nut Cakes',
            'chocolate orange': 'Chocolate Orange Cakes',
            'chocolate orange cake': 'Chocolate Orange Cakes',
            'pinacolada': 'Pina Colada Cakes',
            'zuchinni cake': 'Zucchini Cakes',
            'amarula blackforest': 'Amarula Black Forest Cakes',
            'chocolate chip cake': 'Chocolate Chip Cakes',
            'banana cake': 'Banana Cakes',
            'choc orange b/f': 'Chocolate Orange Black Forest Cakes',
            'coffee cake': 'Coffee Cakes',
            'chocolate fudge cake': 'Chocolate Fudge Cakes',
            'fruit cake': 'Fruit Cakes',
            'rasberry cake': 'Raspberry Cakes',
            'choc mint cake': 'Chocolate Mint Cakes',
            'amarula whiteforest': 'Amarula White Forest Cakes',
            'coconut orange cake': 'Coconut Orange Cakes',
            'ice cream cake': 'Ice Cream Cakes',
            'tiramisu': 'Tiramisu Cakes',
            'cheese cake': 'Cheesecake Cakes',
            'blueberry cake': 'Blueberry Cakes',
            'coconut': 'Coconut Cakes',
            'caramel': 'Caramel Cakes',
        }

        if key in aliases:
            return aliases[key]

        cleaned = re.sub(r'\s+', ' ', text).strip()
        cleaned = cleaned.replace('B/F', 'Black Forest').replace('b/f', 'Black Forest')
        cleaned = cleaned.title()
        cleaned = cleaned.replace('Redvelvet', 'Red Velvet').replace('Blackforest', 'Black Forest')
        cleaned = cleaned.replace('Whiteforest', 'White Forest').replace('Rasberry', 'Raspberry')
        cleaned = cleaned.replace('Zuchinni', 'Zucchini').replace('Pinacolada', 'Pina Colada')
        if not cleaned.endswith('Cake') and not cleaned.endswith('Cakes'):
            cleaned = f'{cleaned} Cakes'
        if cleaned.endswith('Cake'):
            cleaned = f'{cleaned}s'
        return cleaned

    def _matrix_category_for_name(self, raw_name, normalized_title):
        source = f"{str(raw_name or '').lower()} {normalized_title.lower()}"
        raw_key = ' '.join(str(raw_name or '').strip().lower().split())
        normalized_key = ' '.join(str(normalized_title or '').strip().lower().split())

        title_map = {
            'vanilla cakes': 'classic-cakes',
            'strawberry cakes': 'fruit-citrus-cakes',
            'chocolate cakes': 'chocolate-cakes',
            'marble cakes': 'classic-cakes',
            'orange cakes': 'fruit-citrus-cakes',
            'lemon cakes': 'fruit-citrus-cakes',
            'black forest cakes': 'forest-cakes',
            'passion cakes': 'fruit-citrus-cakes',
            'white forest cakes': 'forest-cakes',
            'butterscotch cakes': 'classic-cakes',
            'red velvet cakes': 'classic-cakes',
            'eggless cakes': 'dietary-cakes',
            'diabetic cakes': 'dietary-cakes',
            'cinnamon cherry cakes': 'spiced-nut-cakes',
            'carrot pineapple cakes': 'spiced-nut-cakes',
            'carrot nut cakes': 'spiced-nut-cakes',
            'chocolate orange cakes': 'chocolate-cakes',
            'pina colada cakes': 'fruit-citrus-cakes',
            'zucchini cakes': 'spiced-nut-cakes',
            'amarula black forest cakes': 'premium-cakes',
            'chocolate chip cakes': 'chocolate-cakes',
            'banana cakes': 'fruit-citrus-cakes',
            'chocolate orange black forest cakes': 'premium-cakes',
            'coffee cakes': 'premium-cakes',
            'chocolate fudge cakes': 'chocolate-cakes',
            'fruit cakes': 'fruit-citrus-cakes',
            'raspberry cakes': 'fruit-citrus-cakes',
            'chocolate mint cakes': 'chocolate-cakes',
            'amarula white forest cakes': 'premium-cakes',
            'coconut orange cakes': 'fruit-citrus-cakes',
            'ice cream cakes': 'premium-cakes',
            'tiramisu cakes': 'premium-cakes',
            'cheesecake cakes': 'premium-cakes',
            'blueberry cakes': 'fruit-citrus-cakes',
            'coconut cakes': 'fruit-citrus-cakes',
            'caramel cakes': 'classic-cakes',
        }

        if normalized_key in title_map:
            return title_map[normalized_key]

        explicit_map = {
            'vannila': 'classic-cakes',
            'strawberry': 'fruit-citrus-cakes',
            'chocolate': 'chocolate-cakes',
            'marble': 'classic-cakes',
            'orange': 'fruit-citrus-cakes',
            'lemon': 'fruit-citrus-cakes',
            'blackforest': 'forest-cakes',
            'passion': 'fruit-citrus-cakes',
            'white forest': 'forest-cakes',
            'butter scotch': 'classic-cakes',
            'redvelvet cake': 'classic-cakes',
            'eggless cake': 'dietary-cakes',
            'diabetic cake': 'dietary-cakes',
            'cinnamon cherry': 'spiced-nut-cakes',
            'carrot pineapple': 'spiced-nut-cakes',
            'carrot nut cake': 'spiced-nut-cakes',
            'chocolate orange': 'chocolate-cakes',
            'pinacolada': 'fruit-citrus-cakes',
            'zuchinni cake': 'spiced-nut-cakes',
            'amarula blackforest': 'premium-cakes',
            'chocolate chip cake': 'chocolate-cakes',
            'banana cake': 'fruit-citrus-cakes',
            'choc orange b/f': 'premium-cakes',
            'coffee cake': 'premium-cakes',
            'chocolate fudge cake': 'chocolate-cakes',
            'fruit cake': 'fruit-citrus-cakes',
            'rasberry cake': 'fruit-citrus-cakes',
            'choc mint cake': 'chocolate-cakes',
            'amarula whiteforest': 'premium-cakes',
            'coconut orange cake': 'fruit-citrus-cakes',
            'ice cream cake': 'premium-cakes',
            'tiramisu': 'premium-cakes',
            'cheese cake': 'premium-cakes',
            'blueberry cake': 'fruit-citrus-cakes',
            'coconut': 'fruit-citrus-cakes',
            'caramel': 'classic-cakes',
        }

        if raw_key in explicit_map:
            return explicit_map[raw_key]

        if any(token in source for token in ['eggless', 'diabetic']):
            return 'dietary-cakes'
        if any(token in source for token in ['amarula', 'tiramisu', 'ice cream', 'coffee', 'cheese']):
            return 'premium-cakes'
        if any(token in source for token in ['black forest', 'white forest', 'b/f', 'forest']):
            return 'forest-cakes'
        if any(token in source for token in ['carrot', 'cinnamon', 'zucchini', 'nut']):
            return 'spiced-nut-cakes'
        if any(token in source for token in ['chocolate', 'choc', 'fudge', 'mint']):
            return 'chocolate-cakes'
        if any(token in source for token in ['strawberry', 'lemon', 'orange', 'passion', 'raspberry', 'blueberry', 'fruit', 'banana', 'coconut', 'pineapple', 'pina', 'cherry']):
            return 'fruit-citrus-cakes'
        if any(token in source for token in ['vanilla', 'marble', 'caramel', 'butterscotch', 'red velvet']):
            return 'classic-cakes'
        return 'premium-cakes'

    @transaction.atomic
    def _import_price_matrix_catalog(self, sheet):
        headers, rows = self._sheet_rows(sheet)
        name_col = 'name'
        size_columns = [
            header for header in headers
            if header != name_col and re.search(r'\d+(?:\.\d+)?\s*kg$', header or '', re.IGNORECASE)
        ]
        if not size_columns:
            raise RuntimeError('No weight/size columns found in price-matrix sheet.')

        category_definitions = [
            ('classic-cakes', 'Classic Cakes', 'Core favorites including vanilla, marble, caramel, and butterscotch cakes.'),
            ('chocolate-cakes', 'Chocolate Cakes', 'Chocolate-forward cake collection including fudge, chip, and mint styles.'),
            ('forest-cakes', 'Forest Cakes', 'Black Forest and White Forest cake styles, including specialty twists.'),
            ('fruit-citrus-cakes', 'Fruit & Citrus Cakes', 'Fruit-inspired and citrus cake options like lemon, orange, berry, and tropical blends.'),
            ('spiced-nut-cakes', 'Spiced & Nut Cakes', 'Warm and textured profiles including carrot, cinnamon, zucchini, and nut cakes.'),
            ('premium-cakes', 'Premium Cakes', 'Distinctive premium flavors such as tiramisu, coffee, amarula, ice cream, and cheesecake cakes.'),
            ('dietary-cakes', 'Dietary Cakes', 'Cake options tailored for eggless and diabetic requirements.'),
        ]
        categories = {}
        for idx, (slug, name, description) in enumerate(category_definitions, start=1):
            category, _ = Category.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'description': description,
                    'is_active': True,
                    'order': idx,
                },
            )
            categories[slug] = category

        addon_defs = self._extract_note_addons(sheet)
        addon_objects = []
        for addon_data in addon_defs:
            addon, _ = ProductAddon.objects.update_or_create(
                slug=addon_data['slug'],
                defaults=addon_data,
            )
            addon_objects.append(addon)

        ProductVariant.objects.all().delete()

        product_count = 0
        variant_count = 0
        for row in rows:
            raw_name = row.get(name_col)
            name = str(raw_name).strip() if raw_name is not None else ''
            if not name:
                continue

            prices = []
            for size in size_columns:
                value = row.get(size)
                if value in (None, ''):
                    continue
                price = self._parse_decimal(value, default=None)
                if price is not None:
                    prices.append((size, price))

            if not prices:
                continue

            normalized_title = self._normalized_matrix_title(name)
            category_slug = self._matrix_category_for_name(name, normalized_title)
            category = categories[category_slug]

            base_size, base_price = prices[0]
            product_slug = self._slug(normalized_title)
            if not product_slug:
                continue

            product, _ = Product.objects.update_or_create(
                slug=product_slug,
                defaults={
                    'name': normalized_title,
                    'description': f'{normalized_title} imported from Everest Cakes 2026 price list.',
                    'short_description': normalized_title,
                    'category': category,
                    'base_price': base_price,
                    'stock_quantity': 100,
                    'is_available': True,
                    'is_featured': False,
                    'is_bestseller': False,
                    'is_new': False,
                    'serving_size': '',
                    'min_lead_time': 24,
                    'max_lead_time': 72,
                    'enable_custom_message': True,
                    'max_message_length': 50,
                    'featured_image': '',
                },
            )

            if addon_objects:
                product.addons.set(addon_objects)

            for index, (size_label, price_value) in enumerate(prices, start=1):
                ProductVariant.objects.create(
                    product=product,
                    name=size_label,
                    weight=size_label,
                    price_adjustment=price_value - base_price,
                    stock_quantity=100,
                    is_default=(size_label == base_size),
                    order=index,
                )
                variant_count += 1

            product_count += 1

        self.stdout.write(f'   ✓ Categories imported: {Category.objects.count()}')
        self.stdout.write(f'   ✓ Addons imported: {ProductAddon.objects.count()}')
        self.stdout.write(f'   ✓ Products imported: {product_count}')
        self.stdout.write(f'   ✓ Variants imported: {variant_count}')

    def _get_sheet_by_aliases(self, workbook, aliases, required=True):
        alias_set = {self._normalize_header(name) for name in aliases}
        for sheet_name in workbook.sheetnames:
            normalized = self._normalize_header(sheet_name)
            if normalized in alias_set:
                return workbook[sheet_name]
        if required:
            raise RuntimeError(
                f"Missing required sheet. Expected one of: {', '.join(aliases)}. Found: {', '.join(workbook.sheetnames)}"
            )
        return None

    def _require_any_columns(self, headers, expected_sets, sheet_name):
        available = set(headers)
        for expected in expected_sets:
            if any(column in available for column in expected):
                return
        expected_display = [' or '.join(cols) for cols in expected_sets]
        raise RuntimeError(
            f"Sheet '{sheet_name}' is missing required columns. Expected at least one of: {', '.join(expected_display)}"
        )

    @transaction.atomic
    def _import_categories_sheet(self, sheet):
        headers, rows = self._sheet_rows(sheet)
        self._require_any_columns(headers, [('name',)], sheet.title)

        categories_by_slug = {}
        categories_by_name = {}

        for order_index, row in enumerate(rows, start=1):
            name = str(row.get('name') or '').strip()
            if not name:
                continue

            slug = str(row.get('slug') or '').strip() or slugify(name)
            description = str(row.get('description') or '').strip()
            icon = str(row.get('icon') or '').strip()
            order = self._parse_int(row.get('order'), default=order_index)
            is_active = self._parse_bool(row.get('is_active'), default=True)

            category, _ = Category.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'description': description,
                    'icon': icon,
                    'order': order,
                    'is_active': is_active,
                },
            )

            categories_by_slug[slug.lower()] = category
            categories_by_name[self._normalize_key(name)] = category

        return len(categories_by_slug), categories_by_slug, categories_by_name

    @transaction.atomic
    def _import_addons_sheet(self, sheet):
        headers, rows = self._sheet_rows(sheet)
        self._require_any_columns(headers, [('name',)], sheet.title)

        addons_by_slug = {}
        addons_by_name = {}

        for order_index, row in enumerate(rows, start=1):
            name = str(row.get('name') or '').strip()
            if not name:
                continue

            slug = str(row.get('slug') or '').strip() or slugify(name)
            description = str(row.get('description') or '').strip()
            price = self._parse_decimal(row.get('price'), default=Decimal('0'))
            is_free = self._parse_bool(row.get('is_free'), default=(price == 0))
            max_quantity = self._parse_int(row.get('max_quantity'), default=5)
            order = self._parse_int(row.get('order'), default=order_index)
            is_available = self._parse_bool(row.get('is_available'), default=True)

            addon, _ = ProductAddon.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'description': description,
                    'price': price,
                    'is_free': is_free,
                    'max_quantity': max_quantity,
                    'order': order,
                    'is_available': is_available,
                },
            )

            addons_by_slug[slug.lower()] = addon
            addons_by_name[self._normalize_key(name)] = addon

        return len(addons_by_slug), addons_by_slug, addons_by_name

    @transaction.atomic
    def _import_products_sheet(self, sheet, categories_by_slug, categories_by_name):
        headers, rows = self._sheet_rows(sheet)
        self._require_any_columns(headers, [('name',)], sheet.title)
        self._require_any_columns(headers, [('category', 'category_name', 'category_slug')], sheet.title)

        products_by_slug = {}
        products_by_name = {}

        for row in rows:
            name = str(row.get('name') or '').strip()
            if not name:
                continue

            slug = str(row.get('slug') or '').strip() or slugify(name)

            category_slug = str(row.get('category_slug') or '').strip().lower()
            category_name = self._normalize_key(row.get('category_name') or row.get('category'))
            category = None
            if category_slug:
                category = categories_by_slug.get(category_slug)
            if category is None and category_name:
                category = categories_by_name.get(category_name)

            if category is None:
                raise RuntimeError(
                    f"Product '{name}' references unknown category. Provide category_slug or category/category_name matching categories sheet."
                )

            description = str(row.get('description') or '').strip() or name
            short_description = str(row.get('short_description') or '').strip()
            base_price = self._parse_decimal(row.get('base_price'))
            sale_price_raw = row.get('sale_price')
            sale_price = self._parse_decimal(sale_price_raw, default=None) if sale_price_raw not in (None, '') else None
            cost_price_raw = row.get('cost_price')
            cost_price = self._parse_decimal(cost_price_raw, default=None) if cost_price_raw not in (None, '') else None

            product, _ = Product.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'description': description,
                    'short_description': short_description,
                    'category': category,
                    'base_price': base_price,
                    'sale_price': sale_price,
                    'cost_price': cost_price,
                    'featured_image': str(row.get('featured_image') or ''),
                    'stock_quantity': self._parse_int(row.get('stock_quantity'), default=100),
                    'min_order_quantity': self._parse_int(row.get('min_order_quantity'), default=1),
                    'max_order_quantity': self._parse_int(row.get('max_order_quantity'), default=10),
                    'is_available': self._parse_bool(row.get('is_available'), default=True),
                    'is_featured': self._parse_bool(row.get('is_featured'), default=False),
                    'is_bestseller': self._parse_bool(row.get('is_bestseller'), default=False),
                    'is_new': self._parse_bool(row.get('is_new'), default=False),
                    'serving_size': str(row.get('serving_size') or '').strip(),
                    'min_lead_time': self._parse_int(row.get('min_lead_time'), default=24),
                    'max_lead_time': self._parse_int(row.get('max_lead_time'), default=72),
                    'enable_custom_message': self._parse_bool(row.get('enable_custom_message'), default=True),
                    'max_message_length': self._parse_int(row.get('max_message_length'), default=50),
                },
            )

            products_by_slug[slug.lower()] = product
            products_by_name[self._normalize_key(name)] = product

        return products_by_slug, products_by_name

    @transaction.atomic
    def _import_variants_sheet(self, sheet, products_by_slug, products_by_name):
        headers, rows = self._sheet_rows(sheet)
        self._require_any_columns(headers, [('name', 'variant_name')], sheet.title)
        self._require_any_columns(headers, [('product', 'product_name', 'product_slug')], sheet.title)

        ProductVariant.objects.all().delete()

        created_count = 0
        for order_index, row in enumerate(rows, start=1):
            product_slug = str(row.get('product_slug') or '').strip().lower()
            product_name = self._normalize_key(row.get('product_name') or row.get('product'))

            product = None
            if product_slug:
                product = products_by_slug.get(product_slug)
            if product is None and product_name:
                product = products_by_name.get(product_name)

            if product is None:
                raise RuntimeError(
                    'Variant row references unknown product. Provide product_slug or product/product_name matching products sheet.'
                )

            variant_name = str(row.get('name') or row.get('variant_name') or '').strip()
            if not variant_name:
                continue

            ProductVariant.objects.create(
                product=product,
                name=variant_name,
                weight=str(row.get('weight') or '').strip(),
                price_adjustment=self._parse_decimal(row.get('price_adjustment')),
                stock_quantity=self._parse_int(row.get('stock_quantity'), default=100),
                is_default=self._parse_bool(row.get('is_default'), default=False),
                order=self._parse_int(row.get('order'), default=order_index),
            )
            created_count += 1

        return created_count

    @transaction.atomic
    def _map_product_addons_from_products_sheet(
        self,
        products_sheet,
        products_by_slug,
        products_by_name,
        addons_by_slug,
        addons_by_name,
    ):
        _, rows = self._sheet_rows(products_sheet)

        linked_count = 0
        for row in rows:
            raw_addons = row.get('addons')
            if raw_addons in (None, ''):
                continue

            product_slug = str(row.get('slug') or '').strip().lower()
            product_name = self._normalize_key(row.get('name'))
            product = products_by_slug.get(product_slug) if product_slug else None
            if product is None and product_name:
                product = products_by_name.get(product_name)
            if product is None:
                continue

            addon_tokens = [token.strip() for token in str(raw_addons).split(',') if token and token.strip()]
            if not addon_tokens:
                continue

            product.addons.clear()
            for token in addon_tokens:
                normalized_name = self._normalize_key(token)
                addon = addons_by_slug.get(token.lower()) or addons_by_name.get(normalized_name)
                if addon is None:
                    raise RuntimeError(
                        f"Product '{product.name}' references unknown addon '{token}' in products sheet 'addons' column."
                    )
                product.addons.add(addon)
                linked_count += 1

        return linked_count

    @transaction.atomic
    def _import_product_addons_sheet(
        self,
        sheet,
        products_by_slug,
        products_by_name,
        addons_by_slug,
        addons_by_name,
    ):
        headers, rows = self._sheet_rows(sheet)
        self._require_any_columns(headers, [('product', 'product_name', 'product_slug')], sheet.title)
        self._require_any_columns(headers, [('addon', 'addon_name', 'addon_slug')], sheet.title)

        linked_count = 0
        for row in rows:
            product_slug = str(row.get('product_slug') or '').strip().lower()
            product_name = self._normalize_key(row.get('product_name') or row.get('product'))
            addon_slug = str(row.get('addon_slug') or '').strip().lower()
            addon_name = self._normalize_key(row.get('addon_name') or row.get('addon'))

            product = products_by_slug.get(product_slug) if product_slug else None
            if product is None and product_name:
                product = products_by_name.get(product_name)
            if product is None:
                raise RuntimeError(
                    f"Unknown product in '{sheet.title}' sheet. Row product: {row.get('product') or row.get('product_name') or row.get('product_slug')}"
                )

            addon = addons_by_slug.get(addon_slug) if addon_slug else None
            if addon is None and addon_name:
                addon = addons_by_name.get(addon_name)
            if addon is None:
                raise RuntimeError(
                    f"Unknown addon in '{sheet.title}' sheet. Row addon: {row.get('addon') or row.get('addon_name') or row.get('addon_slug')}"
                )

            product.addons.add(addon)
            linked_count += 1

        return linked_count

    def clear_data(self):
        """Clear existing data"""
        debug_original = settings.DEBUG
        try:
            settings.DEBUG = False
            from django.db import connection

            delete_tables = [
                'cart_cartitem',
                'cart_cart',
                'orders_orderattachment',
                'orders_ordertracking',
                'orders_paymenttransaction',
                'orders_order',
                'orders_enquiry',
                'products_productreview',
                'products_productvariant',
                'products_productattributemapping_available_options',
                'products_productattributemapping',
                'products_product_addons',
                'products_product_tags',
                'products_product',
                'products_productattributeoption',
                'products_productattribute',
                'products_productaddon',
                'products_category',
                'taggit_taggeditem',
                'taggit_tag',
                'core_testimonial',
                'core_herosection',
                'core_featuredcard',
                'core_page',
            ]

            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                existing_tables = {row[0] for row in cursor.fetchall()}
                for table in delete_tables:
                    if table in existing_tables:
                        cursor.execute(f'DELETE FROM "{table}"')
        finally:
            settings.DEBUG = debug_original
        self.stdout.write('   ✓ Data cleared')

    def create_site_settings(self):
        """Create site settings with correct contact information"""
        self.stdout.write('📍 Creating site settings...')
        
        settings, created = SiteSetting.objects.update_or_create(
            pk=1,
            defaults={
                'site_name': 'Everest Cakes',
                'site_description': 'We are Everest Cakes, we are all about cakes. Welcome all. Order celebration cakes from Thika Town.',
                'phone_primary': '+254713311344',
                'phone_secondary': '+254723780139',
                'email': 'info@everestcakes.com',
                'address': 'Thika Town, Kenya',
                'whatsapp_number': '+254713311344',
                'monday_friday': '8:00 AM - 7:00 PM',
                'saturday': '8:00 AM - 7:00 PM',
                'sunday': '10:00 AM - 4:00 PM',
                'facebook_url': 'https://www.facebook.com/everestcakes',
                'instagram_url': 'https://www.instagram.com/everest_cakes/',
                'threads_url': 'https://www.threads.com/@everest_cakes?xmt=AQF02CZKds1olKwKCNyMuUhtTH3yqBsZeaRMpRugUsgCvrQ',
                'twitter_url': '',
                'tiktok_url': '',
                'free_delivery_threshold': Decimal('5000.00'),
                'delivery_fee': Decimal('300.00'),
                'meta_description': 'Everest Cakes in Thika Town. Celebration cakes, custom cake orders, and direct phone or WhatsApp ordering.',
                'meta_keywords': 'Everest Cakes, Thika cakes, custom cakes, birthday cakes, wedding cakes, graduation cakes, cake shop Thika',
            }
        )
        self.stdout.write(f'   ✓ Site settings: {settings.site_name}')

    def create_cms_pages(self):
        """Create CMS pages"""
        self.stdout.write('📄 Creating CMS pages...')
        
        pages = [
            {
                'title': 'About Us',
                'slug': 'about',
                'content': '''<h2>About Everest Cakes</h2>
<p>We are Everest Cakes, and we are all about cakes. Welcome all.</p>

<h3>Where to Find Us</h3>
<p>Visit us in Thika Town or place your order by phone or WhatsApp. We serve customers looking for celebration cakes, custom designs, and cakes for everyday milestones.</p>

<h3>What We Do</h3>
<p>We make cakes for birthdays, weddings, graduations, baby showers, anniversaries, and custom celebrations. Our focus is fresh bakes, clean finishing, and dependable service for every order.</p>

<h3>Contact</h3>
<ul>
<li><strong>Phone:</strong> +254713311344</li>
<li><strong>Alternative Phone:</strong> +254723780139</li>
<li><strong>Threads:</strong> @everest_cakes</li>
<li><strong>Location:</strong> Thika Town, Kenya</li>
</ul>

<p>Tell us what you need, and we will help you plan the right cake for your occasion.</p>''',
                'show_in_footer': True,
                'order': 1,
            },
            {
                'title': 'Privacy Policy',
                'slug': 'privacy',
                'content': '''<h2>Privacy Policy</h2>
<p>At Everest Cakes, we are committed to protecting your privacy and ensuring the security of your personal information.</p>

<h3>Information We Collect</h3>
<p>We collect information you provide when placing orders, including your name, email address, phone number, and delivery address. This information is used solely to process your orders and provide customer support.</p>

<h3>How We Use Your Information</h3>
<ul>
<li>To process and deliver your orders</li>
<li>To communicate with you about your orders</li>
<li>To send promotional offers (with your consent)</li>
<li>To improve our services</li>
</ul>

<h3>Data Security</h3>
<p>We implement appropriate security measures to protect your personal information from unauthorized access, alteration, or disclosure.</p>

<h3>Contact Us</h3>
<p>If you have questions about this privacy policy, please contact us at info@everestcakes.com or call +254713311344.</p>''',
                'show_in_footer': True,
                'order': 2,
            },
            {
                'title': 'Terms of Service',
                'slug': 'terms',
                'content': '''<h2>Terms of Service</h2>
<p>By using Everest Cakes' services, you agree to the following terms:</p>

<h3>Orders</h3>
<p>All orders are subject to availability and confirmation. We reserve the right to refuse or cancel any order.</p>

<h3>Pricing</h3>
<p>Prices are subject to change without notice. The price confirmed at the time of order placement is final.</p>

<h3>Delivery</h3>
<ul>
<li>Delivery times are estimates and may vary due to traffic or weather conditions.</li>
<li>Please inspect your order upon delivery and report any issues immediately.</li>
<li>Delivery fees are non-refundable once delivery is attempted.</li>
</ul>

<h3>Cancellations</h3>
<p>Orders can be cancelled or modified up to 24 hours before the scheduled delivery time. Cancellations within 24 hours may incur a cancellation fee.</p>

<h3>Contact</h3>
<p>For questions about these terms, please contact us at info@everestcakes.com.</p>''',
                'show_in_footer': True,
                'order': 3,
            },
        ]
        
        for page_data in pages:
            page, created = Page.objects.update_or_create(
                slug=page_data['slug'],
                defaults=page_data
            )
            self.stdout.write(f'   ✓ Page: {page.title}')

    def create_categories(self):
        """Create product categories"""
        self.stdout.write('📁 Creating categories...')
        
        categories = [
            {
                'name': 'Wedding Cakes',
                'slug': 'wedding-cakes',
                'description': 'Elegant, bespoke wedding cakes crafted to make your special day unforgettable. From classic tiered designs to modern masterpieces.',
                'icon': 'heart',
                'order': 1,
            },
            {
                'name': 'Birthday Cakes',
                'slug': 'birthday-cakes',
                'description': 'Fun, vibrant birthday cakes for all ages. Custom designs for kids, teens, and adults.',
                'icon': 'cake',
                'order': 2,
            },
            {
                'name': 'Custom Cakes',
                'slug': 'custom-cakes',
                'description': 'Unique, personalized cakes for any occasion. Anniversary, graduation, baby shower, or any celebration.',
                'icon': 'sparkles',
                'order': 3,
            },
            {
                'name': 'Cupcakes & Pastries',
                'slug': 'cupcakes-pastries',
                'description': 'Delicious cupcakes, cookies, and pastries. Perfect for parties, gifts, or personal treats.',
                'icon': 'cookie',
                'order': 4,
            },
            {
                'name': 'Corporate Events',
                'slug': 'corporate-events',
                'description': 'Professional cakes for corporate events, office celebrations, and business milestones.',
                'icon': 'building',
                'order': 5,
            },
            {
                'name': 'Graduation Cakes',
                'slug': 'graduation-cakes',
                'description': 'Celebrate academic achievements with our specially designed graduation cakes.',
                'icon': 'graduation-cap',
                'order': 6,
            },
            {
                'name': 'Baby Shower Cakes',
                'slug': 'baby-shower-cakes',
                'description': 'Adorable cakes for welcoming new arrivals. Gender reveal, baby shower, and christening cakes.',
                'icon': 'baby',
                'order': 7,
            },
            {
                'name': 'Anniversary Cakes',
                'slug': 'anniversary-cakes',
                'description': 'Romantic and elegant anniversary cakes to celebrate your love milestones.',
                'icon': 'heart-handshake',
                'order': 8,
            },
        ]
        
        for cat_data in categories:
            category, created = Category.objects.update_or_create(
                slug=cat_data['slug'],
                defaults=cat_data
            )
            self.stdout.write(f'   ✓ Category: {category.name}')

    def create_attributes(self):
        """Create product attributes and options"""
        self.stdout.write('🎨 Creating attributes...')
        
        # Flavor Attribute
        flavor_attr, _ = ProductAttribute.objects.update_or_create(
            slug='flavor',
            defaults={
                'name': 'Cake Flavor',
                'description': 'Choose your preferred cake flavor',
                'is_required': True,
                'order': 1,
            }
        )
        
        flavors = [
            ('Vanilla', Decimal('0.00')),
            ('Chocolate', Decimal('0.00')),
            ('Red Velvet', Decimal('200.00')),
            ('Black Forest', Decimal('250.00')),
            ('Strawberry', Decimal('200.00')),
            ('Lemon', Decimal('150.00')),
            ('Carrot Cake', Decimal('300.00')),
            ('Marble (Vanilla & Chocolate)', Decimal('100.00')),
            ('Butterscotch', Decimal('200.00')),
            ('Coffee/Mocha', Decimal('250.00')),
            ('Coconut', Decimal('200.00')),
            ('Banana', Decimal('150.00')),
            ('Pineapple', Decimal('200.00')),
            ('Orange', Decimal('150.00')),
            ('German Chocolate', Decimal('350.00')),
        ]
        
        for i, (name, price) in enumerate(flavors):
            ProductAttributeOption.objects.update_or_create(
                attribute=flavor_attr,
                name=name,
                defaults={'price_adjustment': price, 'order': i, 'is_available': True}
            )
        self.stdout.write(f'   ✓ Flavor: {len(flavors)} options')
        
        # Frosting Attribute
        frosting_attr, _ = ProductAttribute.objects.update_or_create(
            slug='frosting',
            defaults={
                'name': 'Frosting Type',
                'description': 'Choose your preferred frosting',
                'is_required': True,
                'order': 2,
            }
        )
        
        frostings = [
            ('Buttercream (Vanilla)', Decimal('0.00')),
            ('Buttercream (Chocolate)', Decimal('0.00')),
            ('Cream Cheese', Decimal('200.00')),
            ('Whipped Cream', Decimal('150.00')),
            ('Chocolate Ganache', Decimal('300.00')),
            ('Fondant', Decimal('500.00')),
            ('Swiss Meringue Buttercream', Decimal('400.00')),
            ('Italian Buttercream', Decimal('400.00')),
        ]
        
        for i, (name, price) in enumerate(frostings):
            ProductAttributeOption.objects.update_or_create(
                attribute=frosting_attr,
                name=name,
                defaults={'price_adjustment': price, 'order': i, 'is_available': True}
            )
        self.stdout.write(f'   ✓ Frosting: {len(frostings)} options')
        
        # Filling Attribute
        filling_attr, _ = ProductAttribute.objects.update_or_create(
            slug='filling',
            defaults={
                'name': 'Filling',
                'description': 'Choose your preferred filling (optional)',
                'is_required': False,
                'order': 3,
            }
        )
        
        fillings = [
            ('No Filling (Just Frosting)', Decimal('0.00')),
            ('Vanilla Custard', Decimal('150.00')),
            ('Chocolate Mousse', Decimal('250.00')),
            ('Strawberry Jam', Decimal('100.00')),
            ('Caramel', Decimal('200.00')),
            ('Nutella', Decimal('300.00')),
            ('Fresh Cream', Decimal('200.00')),
            ('Lemon Curd', Decimal('200.00')),
            ('Pineapple', Decimal('150.00')),
            ('Black Forest Cherry', Decimal('250.00')),
        ]
        
        for i, (name, price) in enumerate(fillings):
            ProductAttributeOption.objects.update_or_create(
                attribute=filling_attr,
                name=name,
                defaults={'price_adjustment': price, 'order': i, 'is_available': True}
            )
        self.stdout.write(f'   ✓ Filling: {len(fillings)} options')
        
        # Shape Attribute
        shape_attr, _ = ProductAttribute.objects.update_or_create(
            slug='shape',
            defaults={
                'name': 'Cake Shape',
                'description': 'Choose the shape of your cake',
                'is_required': True,
                'order': 4,
            }
        )
        
        shapes = [
            ('Round', Decimal('0.00')),
            ('Square', Decimal('200.00')),
            ('Rectangle', Decimal('200.00')),
            ('Heart', Decimal('300.00')),
            ('Number/Letter', Decimal('500.00')),
            ('Custom Shape', Decimal('800.00')),
        ]
        
        for i, (name, price) in enumerate(shapes):
            ProductAttributeOption.objects.update_or_create(
                attribute=shape_attr,
                name=name,
                defaults={'price_adjustment': price, 'order': i, 'is_available': True}
            )
        self.stdout.write(f'   ✓ Shape: {len(shapes)} options')
        
        # Color Theme Attribute
        color_attr, _ = ProductAttribute.objects.update_or_create(
            slug='color-theme',
            defaults={
                'name': 'Color Theme',
                'description': 'Main color for your cake decoration',
                'is_required': True,
                'order': 5,
            }
        )
        
        colors = [
            ('White/Ivory', Decimal('0.00')),
            ('Pink', Decimal('0.00')),
            ('Blue', Decimal('0.00')),
            ('Purple', Decimal('0.00')),
            ('Yellow', Decimal('0.00')),
            ('Green', Decimal('0.00')),
            ('Red', Decimal('0.00')),
            ('Gold', Decimal('100.00')),
            ('Silver', Decimal('100.00')),
            ('Black', Decimal('50.00')),
            ('Rainbow/Multi-color', Decimal('200.00')),
            ('Custom Color', Decimal('150.00')),
        ]
        
        for i, (name, price) in enumerate(colors):
            ProductAttributeOption.objects.update_or_create(
                attribute=color_attr,
                name=name,
                defaults={'price_adjustment': price, 'order': i, 'is_available': True}
            )
        self.stdout.write(f'   ✓ Color Theme: {len(colors)} options')

    def create_addons(self):
        """Create product addons"""
        self.stdout.write('➕ Creating addons...')
        
        addons = [
            {
                'name': 'Birthday Candles (Pack of 10)',
                'slug': 'birthday-candles',
                'description': 'Colorful birthday candles',
                'price': Decimal('50.00'),
                'is_free': False,
                'max_quantity': 5,
                'order': 1,
            },
            {
                'name': 'Number Candles',
                'slug': 'number-candles',
                'description': 'Individual number candles (0-9)',
                'price': Decimal('100.00'),
                'is_free': False,
                'max_quantity': 2,
                'order': 2,
            },
            {
                'name': 'Sparkler Candles',
                'slug': 'sparkler-candles',
                'description': 'Indoor sparkler candles for special celebrations',
                'price': Decimal('200.00'),
                'is_free': False,
                'max_quantity': 3,
                'order': 3,
            },
            {
                'name': 'Message Card',
                'slug': 'message-card',
                'description': 'Beautiful greeting card with your personal message',
                'price': Decimal('150.00'),
                'is_free': False,
                'max_quantity': 1,
                'order': 4,
            },
            {
                'name': 'Gift Box',
                'slug': 'gift-box',
                'description': 'Premium gift box packaging',
                'price': Decimal('300.00'),
                'is_free': False,
                'max_quantity': 1,
                'order': 5,
            },
            {
                'name': 'Edible Image',
                'slug': 'edible-image',
                'description': 'Custom edible image printed on cake (provide your image)',
                'price': Decimal('500.00'),
                'is_free': False,
                'max_quantity': 1,
                'order': 6,
            },
            {
                'name': 'Fresh Flowers Decoration',
                'slug': 'fresh-flowers',
                'description': 'Fresh flower decoration on cake',
                'price': Decimal('600.00'),
                'is_free': False,
                'max_quantity': 1,
                'order': 7,
            },
            {
                'name': 'Figurine/Topper',
                'slug': 'figurine-topper',
                'description': 'Custom cake topper or figurine',
                'price': Decimal('400.00'),
                'is_free': False,
                'max_quantity': 2,
                'order': 8,
            },
            {
                'name': 'Cake Stand (Rental)',
                'slug': 'cake-stand',
                'description': 'Elegant cake stand (deposit required, refundable on return)',
                'price': Decimal('500.00'),
                'is_free': False,
                'max_quantity': 1,
                'order': 9,
            },
            {
                'name': 'Extra Box/Serving Box',
                'slug': 'serving-box',
                'description': 'Additional box for takeaway slices',
                'price': Decimal('100.00'),
                'is_free': False,
                'max_quantity': 10,
                'order': 10,
            },
            {
                'name': 'Knife & Server Set',
                'slug': 'knife-server',
                'description': 'Disposable cake knife and server',
                'price': Decimal('50.00'),
                'is_free': True,
                'max_quantity': 2,
                'order': 11,
            },
            {
                'name': 'Ice Packs',
                'slug': 'ice-packs',
                'description': 'Ice packs for keeping cake fresh during transport',
                'price': Decimal('0.00'),
                'is_free': True,
                'max_quantity': 4,
                'order': 12,
            },
        ]
        
        for addon_data in addons:
            addon, created = ProductAddon.objects.update_or_create(
                slug=addon_data['slug'],
                defaults=addon_data
            )
            self.stdout.write(f'   ✓ Addon: {addon.name} ({addon.display_price})')

    def create_products(self):
        """Create products with variants"""
        self.stdout.write('🎂 Creating products...')
        
        # Get categories
        wedding_cat = Category.objects.get(slug='wedding-cakes')
        birthday_cat = Category.objects.get(slug='birthday-cakes')
        custom_cat = Category.objects.get(slug='custom-cakes')
        cupcakes_cat = Category.objects.get(slug='cupcakes-pastries')
        corporate_cat = Category.objects.get(slug='corporate-events')
        graduation_cat = Category.objects.get(slug='graduation-cakes')
        baby_cat = Category.objects.get(slug='baby-shower-cakes')
        anniversary_cat = Category.objects.get(slug='anniversary-cakes')
        
        # Get attributes
        flavor_attr = ProductAttribute.objects.get(slug='flavor')
        frosting_attr = ProductAttribute.objects.get(slug='frosting')
        
        # Get addons
        all_addons = list(ProductAddon.objects.all())
        
        products_data = [
            # Wedding Cakes
            {
                'name': 'Classic Elegant Wedding Cake',
                'slug': 'classic-elegant-wedding-cake',
                'category': wedding_cat,
                'short_description': 'Timeless elegance for your special day. Classic tiered design with delicate fondant details.',
                'description': '''<h3>A Timeless Masterpiece</h3>
<p>Our Classic Elegant Wedding Cake is the perfect centerpiece for your dream wedding. Featuring multiple tiers of moist, flavorful cake covered in smooth fondant with delicate piped details.</p>

<h4>Features:</h4>
<ul>
<li>Available in 2, 3, or 4 tiers</li>
<li>Choice of cake flavors and fillings</li>
<li>Elegant fondant finish</li>
<li>Customizable decorations</li>
<li>Fresh flower accents available</li>
</ul>

<h4>Perfect For:</h4>
<p>Traditional weddings, church ceremonies, and couples who appreciate classic beauty.</p>''',
                'base_price': Decimal('15000.00'),
                'is_featured': True,
                'is_bestseller': True,
                'serving_size': 'Serves 100-150 guests',
                'min_lead_time': 72,
            },
            {
                'name': 'Modern Minimalist Wedding Cake',
                'slug': 'modern-minimalist-wedding-cake',
                'category': wedding_cat,
                'short_description': 'Clean lines and modern aesthetics for contemporary couples.',
                'description': '''<h3>Contemporary Elegance</h3>
<p>For the modern couple who appreciates clean lines and sophisticated simplicity. This cake features geometric designs, metallic accents, and a sleek finish.</p>

<h4>Features:</h4>
<ul>
<li>Minimalist design with clean edges</li>
<li>Metallic gold or silver accents</li>
<li>Semi-naked or fully frosted options</li>
<li>Available in 2-4 tiers</li>
</ul>''',
                'base_price': Decimal('18000.00'),
                'is_featured': True,
                'serving_size': 'Serves 100-200 guests',
                'min_lead_time': 72,
            },
            {
                'name': 'Rustic Charm Wedding Cake',
                'slug': 'rustic-charm-wedding-cake',
                'category': wedding_cat,
                'short_description': 'Semi-naked cake with fresh flowers and rustic elegance.',
                'description': '''<h3>Natural Beauty</h3>
<p>Perfect for outdoor and barn weddings. This semi-naked cake showcases the beautiful texture of the cake layers, adorned with fresh flowers and greenery.</p>''',
                'base_price': Decimal('12000.00'),
                'is_featured': False,
                'serving_size': 'Serves 80-120 guests',
                'min_lead_time': 48,
            },
            
            # Birthday Cakes
            {
                'name': 'Kids Character Birthday Cake',
                'slug': 'kids-character-birthday-cake',
                'category': birthday_cat,
                'short_description': 'Your child\'s favorite character brought to life in delicious cake form!',
                'description': '''<h3>Make Their Dreams Come True</h3>
<p>From princesses to superheroes, cartoons to animals, we create the perfect character cake for your little one\'s special day.</p>

<h4>Popular Themes:</h4>
<ul>
<li>Disney Princesses & Frozen</li>
<li>Superheroes (Spider-Man, Batman, etc.)</li>
<li>Cartoon Characters (Peppa Pig, Paw Patrol)</li>
<li>Animals & Safari</li>
<li>Cars & Trucks</li>
</ul>''',
                'base_price': Decimal('3500.00'),
                'is_featured': True,
                'is_bestseller': True,
                'serving_size': 'Serves 15-20',
                'min_lead_time': 24,
            },
            {
                'name': 'Teen Trend Birthday Cake',
                'slug': 'teen-trend-birthday-cake',
                'category': birthday_cat,
                'short_description': 'Trendy designs for teenagers - TikTok inspired, gaming, and more!',
                'description': '''<h3>Stay on Trend</h3>
<p>From TikTok viral designs to gaming themes, we create cakes that teens actually want to post about!</p>

<h4>Popular Themes:</h4>
<ul>
<li>TikTok & Social Media</li>
<li>Gaming (Fortnite, Minecraft, Roblox)</li>
<li>K-Pop & Music</li>
<li>Sports & Athletics</li>
</ul>''',
                'base_price': Decimal('4500.00'),
                'is_featured': True,
                'serving_size': 'Serves 20-25',
                'min_lead_time': 24,
            },
            {
                'name': 'Adult Milestone Birthday Cake',
                'slug': 'adult-milestone-birthday-cake',
                'category': birthday_cat,
                'short_description': 'Elegant cakes for milestone birthdays - 21st, 30th, 40th, 50th and beyond!',
                'description': '''<h3>Celebrate in Style</h3>
<p>Milestone birthdays deserve something special. We create sophisticated cakes that reflect the celebrant\'s personality and achievements.</p>''',
                'base_price': Decimal('5500.00'),
                'is_featured': False,
                'serving_size': 'Serves 25-30',
                'min_lead_time': 24,
            },
            {
                'name': 'Number Shaped Birthday Cake',
                'slug': 'number-shaped-birthday-cake',
                'category': birthday_cat,
                'short_description': 'Celebrate the age! Number-shaped cakes in any style.',
                'description': '''<h3>Big Numbers, Big Celebrations</h3>
<p>Make a statement with a cake shaped as the birthday age. Perfect for kids and adults alike!</p>''',
                'base_price': Decimal('4000.00'),
                'is_featured': True,
                'serving_size': 'Serves 15-25',
                'min_lead_time': 24,
            },
            
            # Custom Cakes
            {
                'name': 'Custom Photo Cake',
                'slug': 'custom-photo-cake',
                'category': custom_cat,
                'short_description': 'Your favorite photo printed on a delicious cake!',
                'description': '''<h3>Picture Perfect</h3>
<p>Turn any photo into an edible masterpiece. Perfect for birthdays, anniversaries, memorials, or any special occasion.</p>''',
                'base_price': Decimal('3500.00'),
                'is_featured': True,
                'is_bestseller': True,
                'serving_size': 'Serves 15-20',
                'min_lead_time': 24,
            },
            {
                'name': 'Designer Handbag Cake',
                'slug': 'designer-handbag-cake',
                'category': custom_cat,
                'short_description': 'Stunning designer-inspired handbag cakes for fashion lovers.',
                'description': '''<h3>Fashion Forward</h3>
<p>Our designer handbag cakes are showstoppers! We recreate iconic designs in delicious cake form.</p>''',
                'base_price': Decimal('6500.00'),
                'is_featured': True,
                'serving_size': 'Serves 20-30',
                'min_lead_time': 48,
            },
            {
                'name': 'Alcohol Bottle Cake',
                'slug': 'alcohol-bottle-cake',
                'category': custom_cat,
                'short_description': 'Realistic alcohol bottle cakes for adult celebrations.',
                'description': '''<h3>Cheers!</h3>
<p>Perfect for bachelor parties, 21st birthdays, or any adult celebration. We create realistic alcohol bottle cakes that look just like the real thing.</p>''',
                'base_price': Decimal('5000.00'),
                'is_featured': False,
                'serving_size': 'Serves 20-25',
                'min_lead_time': 24,
            },
            
            # Cupcakes & Pastries
            {
                'name': 'Classic Cupcakes (Dozen)',
                'slug': 'classic-cupcakes-dozen',
                'category': cupcakes_cat,
                'short_description': 'Delicious cupcakes in your choice of flavor. Perfect for parties!',
                'description': '''<h3>Bite-Sized Delights</h3>
<p>Our cupcakes are made fresh daily with premium ingredients. Available in a variety of flavors and decorated to match your theme.</p>

<h4>Flavors Available:</h4>
<p>Vanilla, Chocolate, Red Velvet, Strawberry, Lemon, and more!</p>''',
                'base_price': Decimal('1200.00'),
                'is_featured': True,
                'is_bestseller': True,
                'serving_size': '12 cupcakes',
                'min_lead_time': 12,
            },
            {
                'name': 'Gourmet Cupcakes (Box of 6)',
                'slug': 'gourmet-cupcakes-box',
                'category': cupcakes_cat,
                'short_description': 'Premium gourmet cupcakes with exotic flavors and elegant decorations.',
                'description': '''<h3>Premium Cupcakes</h3>
<p>Our gourmet cupcakes feature premium ingredients and exotic flavors for a truly indulgent experience.</p>''',
                'base_price': Decimal('900.00'),
                'is_featured': False,
                'serving_size': '6 cupcakes',
                'min_lead_time': 12,
            },
            {
                'name': 'Cookies (Box of 20)',
                'slug': 'cookies-box',
                'category': cupcakes_cat,
                'short_description': 'Fresh baked cookies in various flavors. Perfect for gifts!',
                'description': '''<h3>Homestyle Cookies</h3>
<p>Soft, chewy cookies made with love. Available in chocolate chip, oatmeal raisin, and more.</p>''',
                'base_price': Decimal('800.00'),
                'is_featured': False,
                'serving_size': '20 cookies',
                'min_lead_time': 12,
            },
            
            # Corporate Events
            {
                'name': 'Corporate Logo Cake',
                'slug': 'corporate-logo-cake',
                'category': corporate_cat,
                'short_description': 'Your company logo beautifully recreated on a delicious cake.',
                'description': '''<h3>Brand Your Celebration</h3>
<p>Perfect for product launches, company anniversaries, and corporate milestones. We recreate your company logo with precision on a cake that feeds the whole team.</p>''',
                'base_price': Decimal('6000.00'),
                'is_featured': True,
                'serving_size': 'Serves 30-50',
                'min_lead_time': 48,
            },
            {
                'name': 'Office Celebration Cake',
                'slug': 'office-celebration-cake',
                'category': corporate_cat,
                'short_description': 'Professional sheet cakes for office parties and celebrations.',
                'description': '''<h3>Feed the Team</h3>
<p>Our office celebration cakes are designed to serve large groups efficiently while still looking professional and tasting delicious.</p>''',
                'base_price': Decimal('4000.00'),
                'is_featured': False,
                'serving_size': 'Serves 40-50',
                'min_lead_time': 24,
            },
            
            # Graduation Cakes
            {
                'name': 'Graduation Achievement Cake',
                'slug': 'graduation-achievement-cake',
                'category': graduation_cat,
                'short_description': 'Celebrate academic success with a custom graduation cake!',
                'description': '''<h3>Class of Excellence</h3>
<p>From kindergarten to university, we create graduation cakes that honor academic achievements. Complete with cap, diploma, and school colors!</p>''',
                'base_price': Decimal('4000.00'),
                'is_featured': True,
                'serving_size': 'Serves 20-30',
                'min_lead_time': 24,
            },
            {
                'name': 'Graduation Sheet Cake',
                'slug': 'graduation-sheet-cake',
                'category': graduation_cat,
                'short_description': 'Large sheet cake perfect for graduation parties.',
                'description': '''<h3>Party Ready</h3>
<p>Our graduation sheet cakes feed a crowd and can be customized with the graduate\'s name, school, and year.</p>''',
                'base_price': Decimal('3500.00'),
                'is_featured': False,
                'serving_size': 'Serves 30-40',
                'min_lead_time': 24,
            },
            
            # Baby Shower Cakes
            {
                'name': 'Baby Shower Delight',
                'slug': 'baby-shower-delight',
                'category': baby_cat,
                'short_description': 'Adorable cakes for welcoming little ones into the world.',
                'description': '''<h3>Sweet Beginnings</h3>
<p>Our baby shower cakes are designed with love and care. From cute baby themes to elegant designs, we create the perfect centerpiece for your celebration.</p>''',
                'base_price': Decimal('4000.00'),
                'is_featured': True,
                'serving_size': 'Serves 20-25',
                'min_lead_time': 24,
            },
            {
                'name': 'Gender Reveal Cake',
                'slug': 'gender-reveal-cake',
                'category': baby_cat,
                'short_description': 'Surprise! Pink or blue inside to reveal your baby\'s gender.',
                'description': '''<h3>The Big Reveal</h3>
<p>Make your gender reveal party unforgettable! The outside is neutral, but cut inside to reveal pink or blue. We keep the secret until the big moment!</p>''',
                'base_price': Decimal('4500.00'),
                'is_featured': True,
                'is_bestseller': True,
                'serving_size': 'Serves 20-25',
                'min_lead_time': 24,
            },
            {
                'name': 'Christening Cake',
                'slug': 'christening-cake',
                'category': baby_cat,
                'short_description': 'Beautiful cakes for your baby\'s special blessing day.',
                'description': '''<h3>Blessed Beginnings</h3>
<p>Elegant christening cakes with crosses, doves, and soft pastel colors. Perfect for the special day.</p>''',
                'base_price': Decimal('3500.00'),
                'is_featured': False,
                'serving_size': 'Serves 20-30',
                'min_lead_time': 24,
            },
            
            # Anniversary Cakes
            {
                'name': 'Romantic Anniversary Cake',
                'slug': 'romantic-anniversary-cake',
                'category': anniversary_cat,
                'short_description': 'Celebrate your love with a beautiful anniversary cake.',
                'description': '''<h3>Love Stories</h3>
<p>From first anniversaries to golden celebrations, we create romantic cakes that tell your love story. Hearts, flowers, and elegant designs.</p>''',
                'base_price': Decimal('4500.00'),
                'is_featured': True,
                'serving_size': 'Serves 15-25',
                'min_lead_time': 24,
            },
            {
                'name': 'Milestone Anniversary Cake',
                'slug': 'milestone-anniversary-cake',
                'category': anniversary_cat,
                'short_description': 'Silver, Golden, Diamond anniversary cakes for major milestones.',
                'description': '''<h3>Golden Moments</h3>
<p>25th, 50th, or 60th anniversary? We create stunning milestone cakes with elegant silver or gold accents.</p>''',
                'base_price': Decimal('6000.00'),
                'is_featured': False,
                'serving_size': 'Serves 30-50',
                'min_lead_time': 48,
            },
        ]
        
        # Create products
        for product_data in products_data:
            addons = product_data.pop('addons', all_addons)
            category = product_data.pop('category')
            
            product, created = Product.objects.update_or_create(
                slug=product_data['slug'],
                defaults={**product_data, 'category': category, 'stock_quantity': 100}
            )
            
            # Add attributes
            ProductAttributeMapping.objects.get_or_create(
                product=product,
                attribute=flavor_attr,
                defaults={'is_required': True, 'order': 1}
            )
            ProductAttributeMapping.objects.get_or_create(
                product=product,
                attribute=frosting_attr,
                defaults={'is_required': True, 'order': 2}
            )
            
            # Add addons
            for addon in addons:
                product.addons.add(addon)
            
            # Create variants (sizes)
            variants_data = [
                {'name': 'Small', 'weight': '1kg', 'price_adjustment': Decimal('0.00'), 'stock_quantity': 100, 'is_default': True, 'order': 1},
                {'name': 'Medium', 'weight': '2kg', 'price_adjustment': Decimal('1500.00'), 'stock_quantity': 100, 'is_default': False, 'order': 2},
                {'name': 'Large', 'weight': '3kg', 'price_adjustment': Decimal('3000.00'), 'stock_quantity': 100, 'is_default': False, 'order': 3},
                {'name': 'Extra Large', 'weight': '4kg', 'price_adjustment': Decimal('4500.00'), 'stock_quantity': 100, 'is_default': False, 'order': 4},
            ]
            
            for variant_data in variants_data:
                ProductVariant.objects.update_or_create(
                    product=product,
                    name=variant_data['name'],
                    defaults=variant_data
                )
            
            status = '★' if product.is_featured else ' '
            bestseller = '🥇' if product.is_bestseller else ''
            self.stdout.write(f'   {status} {product.name} {bestseller} (KSh {product.base_price})')

    def create_testimonials(self):
        """Create testimonials"""
        self.stdout.write('💬 Creating testimonials...')

        seeded_names = [
            'Mary Wanjiku',
            'James Kamau',
            'Grace Njeri',
            'Peter Mwangi',
            'Lydia Muthoni',
            'Samuel Kariuki',
            'Faith Wambui',
            'Daniel Ochieng',
        ]
        deleted_count, _ = Testimonial.objects.filter(customer_name__in=seeded_names).delete()
        self.stdout.write(f'   ✓ Removed {deleted_count} seeded testimonial records')
        self.stdout.write('   ✓ No testimonials are seeded by default. Add verified customer feedback from admin.')

    def create_hero_sections(self):
        """Create hero sections"""
        self.stdout.write('🖼️  Creating hero sections...')

        HeroSection.objects.filter(title='Treat yourself to decadent moments.').delete()
        
        hero_data = [
            {
                'title': 'Everest Cakes for every celebration.',
                'subtitle': 'Thika Town Cake Studio',
                'description': 'Order custom cakes, celebration cakes, and event cakes directly from Everest Cakes in Thika Town.',
                'background_type': 'image',
                'title_animation': 'fade',
                'content_animation': 'slide',
                'cta_text': 'Explore Menu',
                'cta_link': '/products/',
                'secondary_cta_text': 'Contact Us',
                'secondary_cta_link': '/contact/',
                'is_active': True,
                'order': 1,
            },
        ]
        
        for data in hero_data:
            hero, created = HeroSection.objects.update_or_create(
                title=data['title'],
                defaults=data
            )
            self.stdout.write(f'   ✓ Hero: {hero.title[:50]}...')

    def create_featured_cards(self):
        """Create featured cards"""
        self.stdout.write('🃏 Creating featured cards...')

        FeaturedCard.objects.filter(title__in=[
            'Personal Customizations',
            'Unparalleled Expertise',
            'Timely Delivery',
        ]).delete()
        
        cards = [
            {
                'title': 'Custom Orders',
                'subtitle': 'Made For Your Event',
                'description': 'Share your preferred size, flavor, colors, and message and we will guide you to the right cake.',
                'icon': 'palette',
                'animation_type': 'hover-lift',
                'show_on_homepage': True,
                'order': 1,
            },
            {
                'title': 'Celebration Cakes',
                'subtitle': 'Birthdays To Weddings',
                'description': 'We bake for birthdays, graduations, baby showers, anniversaries, weddings, and other special occasions.',
                'icon': 'award',
                'animation_type': 'hover-lift',
                'show_on_homepage': True,
                'order': 2,
            },
            {
                'title': 'Direct Ordering',
                'subtitle': 'Call Or WhatsApp',
                'description': 'Reach us quickly by phone or WhatsApp to confirm availability, timing, pickup, or delivery.',
                'icon': 'truck',
                'animation_type': 'hover-lift',
                'show_on_homepage': True,
                'order': 3,
            },
        ]
        
        for card_data in cards:
            card, created = FeaturedCard.objects.update_or_create(
                title=card_data['title'],
                defaults=card_data
            )
            self.stdout.write(f'   ✓ Card: {card.title}')
